"""联系人管理API控制器"""
from io import BytesIO

from flask import Blueprint, request, jsonify, send_file
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook, load_workbook

from exts import db
from models import ContactMethodModel, UserModel

# 创建蓝图
bp = Blueprint('contacts', __name__)

ALLOWED_CONTACT_TYPES = {'phone', 'email', 'social', 'address', 'other'}


def _normalize_contact_methods(raw_methods):
    """将请求中的联系方式列表进行标准化和过滤"""
    methods = []
    for item in raw_methods or []:
        contact_type = (item.get('type') or '').strip().lower()
        value = (item.get('value') or '').strip()
        label = (item.get('label') or '').strip()

        if not contact_type or not value:
            continue
        if contact_type not in ALLOWED_CONTACT_TYPES:
            continue

        methods.append({
            'type': contact_type,
            'value': value,
            'label': label
        })
    return methods


def _apply_contact_methods(user, methods):
    """替换某个联系人的联系方式列表"""
    user.contact_methods.clear()
    for method in methods:
        user.contact_methods.append(
            ContactMethodModel(
                contact_type=method['type'],
                value=method['value'],
                label=method.get('label') or ''
            )
        )


@bp.route('/users', methods=['GET'])
def get_users():
    """
    获取联系人列表（支持搜索）

    查询参数:
        keyword: 搜索关键词，用于匹配姓名或电话号码

    返回:
        JSON响应，包含联系人列表
    """
    keyword = request.args.get('keyword', '').strip()
    query = UserModel.query

    if keyword:
        query = query.filter(
            (UserModel.username.like(f"%{keyword}%")) |
            (UserModel.phone_number.like(f"%{keyword}%"))
        )

    # 收藏优先排序
    users = query.order_by(UserModel.is_favorite.desc(), UserModel.id.asc()).all()
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': [user.to_dict() for user in users]
    })


@bp.route('/users', methods=['POST'])
def add_user():
    """
    添加新联系人

    请求体:
        JSON对象，包含username(必填), phone_number/address(可选), is_favorite(可选),
        contact_methods(可选列表)

    返回:
        JSON响应，包含添加结果和新创建的联系人信息
    """
    data = request.get_json() or {}

    # 验证必要字段
    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'code': 400, 'message': '缺少必要参数: username'}), 400

    contact_methods = _normalize_contact_methods(data.get('contact_methods'))
    phone_number = (data.get('phone_number') or '').strip() or None
    address = (data.get('address') or '').strip() or None
    is_favorite = bool(data.get('is_favorite', False))

    # 没有任何联系方式则拒绝
    if not contact_methods and not phone_number and not address:
        return jsonify({'code': 400, 'message': '至少提供一种联系方式'}), 400

    try:
        # 如果未指定主电话/地址，使用第一条同类联系方式作为主信息
        if not phone_number:
            for method in contact_methods:
                if method['type'] == 'phone':
                    phone_number = method['value']
                    break
        if not address:
            for method in contact_methods:
                if method['type'] == 'address':
                    address = method['value']
                    break

        user = UserModel(
            username=username,
            phone_number=phone_number,
            address=address,
            is_favorite=is_favorite
        )
        _apply_contact_methods(user, contact_methods)

        db.session.add(user)
        db.session.commit()

        return jsonify({
            'code': 201,
            'message': '添加成功',
            'data': user.to_dict()
        }), 201

    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'code': 400,
            'message': '数据库约束错误'
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'服务器错误: {str(e)}'
        }), 500


@bp.route('/users/<int:user_id>', methods=['GET'])
def get_user(user_id):
    """
    获取单个联系人详情

    参数:
        user_id: 联系人ID

    返回:
        JSON响应，包含联系人详情
    """
    user = UserModel.query.get(user_id)
    if not user:
        return jsonify({
            'code': 404,
            'message': '联系人不存在'
        }), 404

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': user.to_dict()
    })


@bp.route('/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """
    更新联系人信息

    参数:
        user_id: 联系人ID

    请求体:
        JSON对象，包含需要更新的字段（username, phone_number, address, is_favorite, contact_methods）

    返回:
        JSON响应，包含更新结果和更新后的联系人信息
    """
    user = UserModel.query.get(user_id)
    if not user:
        return jsonify({
            'code': 404,
            'message': '联系人不存在'
        }), 404

    data = request.get_json() or {}

    # 更新字段（只更新提供的字段）
    if 'username' in data:
        user.username = (data['username'] or '').strip()
    if 'phone_number' in data:
        user.phone_number = (data['phone_number'] or '').strip() or None
    if 'address' in data:
        user.address = (data['address'] or '').strip() or None
    if 'is_favorite' in data:
        user.is_favorite = bool(data['is_favorite'])
    if 'contact_methods' in data:
        normalized = _normalize_contact_methods(data.get('contact_methods'))
        if not normalized and not user.phone_number and not user.address:
            return jsonify({'code': 400, 'message': '至少提供一种联系方式'}), 400
        _apply_contact_methods(user, normalized)

    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '修改成功',
            'data': user.to_dict()
        })
    except IntegrityError:
        db.session.rollback()
        return jsonify({
            'code': 400,
            'message': '数据库约束错误'
        }), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'服务器错误: {str(e)}'
        }), 500


@bp.route('/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """
    删除联系人

    参数:
        user_id: 联系人ID

    返回:
        JSON响应，包含删除结果
    """
    user = UserModel.query.get(user_id)
    if not user:
        return jsonify({
            'code': 404,
            'message': '联系人不存在'
        }), 404

    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '删除成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'message': f'删除失败: {str(e)}'
        }), 500


@bp.route('/users/export', methods=['GET'])
def export_users():
    """导出所有联系人到Excel（使用openpyxl，避免pandas依赖）"""
    users = UserModel.query.order_by(UserModel.is_favorite.desc(), UserModel.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '联系人'

    headers = ['姓名', '是否收藏', '电话', '邮箱', '社交账号', '地址', '其他']
    ws.append(headers)

    for user in users:
        grouped = {'phone': [], 'email': [], 'social': [], 'address': [], 'other': []}
        for method in user.contact_methods:
            grouped.get(method.contact_type, grouped['other']).append(method.value)

        if user.phone_number and user.phone_number not in grouped['phone']:
            grouped['phone'].insert(0, user.phone_number)
        if user.address and user.address not in grouped['address']:
            grouped['address'].insert(0, user.address)

        ws.append([
            user.username,
            '是' if user.is_favorite else '否',
            '\n'.join(grouped['phone']),
            '\n'.join(grouped['email']),
            '\n'.join(grouped['social']),
            '\n'.join(grouped['address']),
            '\n'.join(grouped['other']),
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='contacts.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/users/import', methods=['POST'])
def import_users():
    """从Excel导入联系人（使用openpyxl，避免pandas依赖）"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传Excel文件'}), 400

    file = request.files['file']
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'code': 400, 'message': f'解析Excel失败: {str(e)}'}), 400

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if '姓名' not in headers:
        return jsonify({'code': 400, 'message': 'Excel缺少“姓名”列'}), 400

    col_index = {name: idx for idx, name in enumerate(headers)}

    def parse_column(row_values, col_name):
        value = row_values[col_index.get(col_name, -1)] if col_name in col_index else None
        raw = str(value or '').strip()
        if not raw:
            return []
        parts = [part.strip() for part in raw.replace(';', '\n').replace(',', '\n').splitlines()]
        return [p for p in parts if p]

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        username = str(row[col_index['姓名']] or '').strip()
        if not username:
            continue

        is_favorite_raw = str(row[col_index.get('是否收藏', -1)] or '').strip()
        is_favorite = is_favorite_raw in {'是', '1', 'true', 'True', 'Y', 'yes'}

        phones = parse_column(row, '电话')
        emails = parse_column(row, '邮箱')
        socials = parse_column(row, '社交账号')
        addresses = parse_column(row, '地址')
        others = parse_column(row, '其他')

        contact_methods = []
        contact_methods += [{'type': 'phone', 'value': p} for p in phones]
        contact_methods += [{'type': 'email', 'value': e} for e in emails]
        contact_methods += [{'type': 'social', 'value': s} for s in socials]
        contact_methods += [{'type': 'address', 'value': a} for a in addresses]
        contact_methods += [{'type': 'other', 'value': o} for o in others]

        primary_phone = phones[0] if phones else None
        primary_address = addresses[0] if addresses else None

        if not contact_methods and not primary_phone and not primary_address:
            continue

        user = UserModel(
            username=username,
            phone_number=primary_phone,
            address=primary_address,
            is_favorite=is_favorite
        )
        _apply_contact_methods(user, contact_methods)
        db.session.add(user)
        created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500

    return jsonify({'code': 200, 'message': f'导入成功，新增 {created} 条联系人'}), 200